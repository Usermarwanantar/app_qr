import streamlit as st
import qrcode
from PIL import Image, ImageDraw, ImageFont
import io
import os
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker
from openpyxl.utils.units import pixels_to_EMU
import pandas as pd

# Configuration de la page
st.set_page_config(page_title="Générateur de QR Code - OneDrive", page_icon="📤", layout="centered")

# En-tête
st.markdown(
    "<h1 style='text-align: center; color: #2E86C1;'>📤 Générateur de QR Code pour fichiers OneDrive</h1>",
    unsafe_allow_html=True
)

# Instructions
with st.expander("ℹ️ Instructions (clique ici)", expanded=True):
    st.markdown("""
        <div style='background-color: #F0F8FF; padding: 15px; border-radius: 10px;'>
        1️⃣ Téléverse ton fichier sur <a href='https://onedrive.live.com/' target='_blank'>OneDrive</a><br>
        2️⃣ Active l'option de <strong>partage</strong> (Toute personne avec le lien...)<br>
        3️⃣ Colle le lien ci-dessous 👇
        </div>
        """, unsafe_allow_html=True)

# Formulaire utilisateur
st.markdown("---")
st.markdown("<h4 style='color:#117A65;'>📝 Formulaire QR Code</h4>", unsafe_allow_html=True)

project_name = st.text_input("🏷️ Nom du projet")
dtr = st.text_input("🗂️ DTR")
shared_link = st.text_input("🔗 Lien OneDrive")
file_type = st.selectbox("📂 Type de fichier", ["Plugmaps", "Wirliste"])
file_title = st.text_input("✏️ Titre du document")

# Fichiers
qr_image_folder = "qr_images"
os.makedirs(qr_image_folder, exist_ok=True)

# Affichage de l'historique Excel
history_file = "historique1_qr.xlsx"
if os.path.exists(history_file):
    try:
        df_history = pd.read_excel(history_file)
        st.markdown("<h5 style='color:#2874A6;'>📑 Historique des QR Codes générés</h5>", unsafe_allow_html=True)
        # Affichage personnalisé avec images QR
        headers = ["Nom du projet", "DTR", "Titre", "Type", "Lien partagé", "QR Code"]
        cols = st.columns([2, 1, 2, 1, 3, 1])
        for i, h in enumerate(headers):
            cols[i].markdown(f"**{h}**")
        for idx, row in df_history.iterrows():
            cols = st.columns([2, 1, 2, 1, 3, 1])
            cols[0].write(row.get("Nom du projet", ""))
            cols[1].write(row.get("DTR", ""))
            cols[2].write(row.get("Titre", ""))
            cols[3].write(row.get("Type", ""))
            cols[4].write(row.get("Lien partagé", ""))
            qr_filename = f"{str(row.get('Titre', '')).replace(' ', '_')}_QR.png"
            qr_path = os.path.join(qr_image_folder, qr_filename)
            if os.path.exists(qr_path):
                cols[5].image(qr_path, width=120)
            else:
                cols[5].write("(Non trouvé)")
        # Bouton de téléchargement Excel
        with open(history_file, "rb") as f:
            st.download_button(
                label="⬇️ Télécharger l'historique Excel",
                data=f,
                file_name=history_file,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        # Sélecteur de titre et affichage du QR code correspondant
        titres = df_history["Titre"].dropna().unique().tolist()
        if titres:
            st.markdown("<h5 style='color:#239B56;'>🔍 Afficher un QR Code par titre</h5>", unsafe_allow_html=True)
            selected_titre = st.selectbox("Choisis un titre pour voir son QR code :", titres)
            if selected_titre:
                qr_filename = f"{str(selected_titre).replace(' ', '_')}_QR.png"
                qr_path = os.path.join(qr_image_folder, qr_filename)
                if os.path.exists(qr_path):
                    st.image(qr_path, caption=f"QR Code pour : {selected_titre}", width=300)
                else:
                    st.warning("QR code non trouvé pour ce titre.")
    except Exception as e:
        st.warning(f"Erreur lors de la lecture de l'historique : {e}")
else:
    st.info("Aucun historique trouvé. Génère un QR Code pour commencer à enregistrer l'historique.")

if shared_link and file_title and file_type and project_name and dtr:
    if st.button("📷 Générer le QR Code"):

        # Génération QR enrichi
        qr = qrcode.make(shared_link)
        qr = qr.resize((300, 300))
        width, height = 400, 420
        qr_image = Image.new("RGB", (width, height), "white")
        draw = ImageDraw.Draw(qr_image)

        try:
            font_title = ImageFont.truetype("arial.ttf", 18)
            font_type = ImageFont.truetype("arial.ttf", 16)
        except:
            font_title = font_type = None

        type_text = f"[📂 {file_type} File]"
        bbox_type = draw.textbbox((0, 0), type_text, font=font_type)
        w_type = bbox_type[2] - bbox_type[0]
        draw.text(((width - w_type) / 2, 10), type_text, fill="black", font=font_type)

        qr_image.paste(qr, (50, 40))

        bbox_title = draw.textbbox((0, 0), file_title, font=font_title)
        w_title = bbox_title[2] - bbox_title[0]
        draw.text(((width - w_title) / 2, 360), file_title, fill="black", font=font_title)

        # Affichage
        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown("<h4 style='color:#884EA0;'>🎉 Résultat généré :</h4>", unsafe_allow_html=True)
        st.image(qr_image, caption="QR Code prêt à être partagé", use_column_width=False)

        # Sauvegarde image
        qr_filename = f"{file_title.replace(' ', '_')}_QR.png"
        qr_path = os.path.join(qr_image_folder, qr_filename)
        qr_image.save(qr_path)

        with open(qr_path, "rb") as f:
            st.download_button(
                label="📥 Télécharger le QR Code",
                data=f,
                file_name=qr_filename,
                mime="image/png"
            )

        # Création Excel si nécessaire
        if not os.path.exists(history_file):
            wb = Workbook()
            ws = wb.active
            ws.append(["Nom du projet", "DTR", "Titre", "Type", "Lien partagé", "QR Code"])
            wb.save(history_file)

        # Chargement Excel
        wb = load_workbook(history_file)
        ws = wb.active
        next_row = ws.max_row + 1

        # Données texte
        ws.cell(row=next_row, column=1).value = project_name
        ws.cell(row=next_row, column=2).value = dtr
        ws.cell(row=next_row, column=3).value = file_title
        ws.cell(row=next_row, column=4).value = file_type
        ws.cell(row=next_row, column=5).value = shared_link

        # Image QR centrée dans F{next_row}
        img = XLImage(qr_path)
        img.width = 100
        img.height = 100
        cell_ref = f"F{next_row}"
        ws.add_image(img, cell_ref)
        ws.row_dimensions[next_row].height = 120

        # Mise en forme entêtes
        headers = ["Nom du projet", "DTR", "Titre", "Type", "Lien partagé", "QR Code"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Largeur des colonnes
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 50
        ws.column_dimensions["F"].width = 18

        # Centrage du texte
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
            for idx, cell in enumerate(row):
                if idx == 4:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        try:
            wb.save(history_file)
            st.success("✅ QR Code bien enregistré avec image centrée dans Excel.")
            st.rerun()
        except PermissionError:
            st.error("❌ Fichier Excel ouvert. Ferme-le puis réessaie.")
else:
    st.warning("🟠 Remplis tous les champs pour générer le QR Code.")
